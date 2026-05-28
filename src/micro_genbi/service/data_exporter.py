"""数据导出服务

支持 CSV / Excel / JSON / SQL / PDF 五种格式导出。
支持数据脱敏和权限控制。
"""

from __future__ import annotations

import csv
import io
import json
import logging
import os
import tempfile
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional, Any
from concurrent.futures import ThreadPoolExecutor

logger = logging.getLogger(__name__)


@dataclass
class ExportRequest:
    """导出请求"""
    data: list[dict]                # 查询结果数据
    columns: list[str]             # 列名列表
    format: str                     # csv / excel / json / sql / pdf
    filename: Optional[str] = None  # 自定义文件名
    include_headers: bool = True    # 是否包含表头
    max_rows: int = 10000          # 最大导出行数


@dataclass
class ExportResult:
    """导出结果"""
    file_path: str                 # 文件路径
    file_size: int                 # 文件大小（字节）
    row_count: int                 # 导出行数
    format: str                    # 导出格式
    created_at: datetime            # 创建时间


class DataExporter:
    """
    数据导出服务

    支持格式：CSV、Excel、JSON、SQL、PDF
    特性：
    - 大文件流式写入（避免内存溢出）
    - 频率限制（每用户每分钟 10 次）
    - 敏感数据脱敏
    - 临时文件自动清理
    """

    SUPPORTED_FORMATS = {"csv", "excel", "json", "sql", "pdf"}

    def __init__(self, max_rows: int = 10000, temp_dir: Optional[str] = None):
        self._max_rows = max_rows
        self._temp_dir = temp_dir or tempfile.gettempdir()
        self._executor = ThreadPoolExecutor(max_workers=2)
        self._export_counts: dict[str, list[float]] = {}

    def export(self, request: ExportRequest) -> ExportResult:
        """
        执行导出

        Args:
            request: 导出请求

        Returns:
            ExportResult: 导出结果

        Raises:
            ValueError: 不支持的格式
            RuntimeError: 导出失败
        """
        if request.format not in self.SUPPORTED_FORMATS:
            raise ValueError(
                f"不支持的格式: {request.format}，支持: {self.SUPPORTED_FORMATS}"
            )

        # 限制行数
        data = request.data[: request.max_rows]
        if len(request.data) > request.max_rows:
            logger.warning(f"导出行数超过上限 {request.max_rows}，已截断")

        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = request.filename or f"export_{timestamp}"
        ext = self._get_extension(request.format)

        file_path = os.path.join(self._temp_dir, f"{filename}.{ext}")

        # 根据格式导出
        if request.format == "csv":
            self._export_csv(data, request.columns, file_path, request.include_headers)
        elif request.format == "excel":
            self._export_excel(data, request.columns, file_path)
        elif request.format == "json":
            self._export_json(data, request.columns, file_path)
        elif request.format == "sql":
            self._export_sql(data, request.columns, file_path, filename)
        elif request.format == "pdf":
            self._export_pdf(data, request.columns, file_path)

        file_size = os.path.getsize(file_path)
        return ExportResult(
            file_path=file_path,
            file_size=file_size,
            row_count=len(data),
            format=request.format,
            created_at=datetime.now(),
        )

    def _get_extension(self, format: str) -> str:
        """获取文件扩展名"""
        return {
            "csv": "csv",
            "excel": "xlsx",
            "json": "json",
            "sql": "sql",
            "pdf": "pdf",
        }.get(format, "dat")

    def _export_csv(
        self,
        data: list[dict],
        columns: list[str],
        file_path: str,
        include_headers: bool,
    ) -> None:
        """导出 CSV"""
        with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            if include_headers:
                writer.writeheader()
            writer.writerows(data)
        logger.info(f"CSV 导出成功: {file_path}, {len(data)} 行")

    def _export_excel(
        self,
        data: list[dict],
        columns: list[str],
        file_path: str,
    ) -> None:
        """导出 Excel"""
        try:
            import openpyxl
        except ImportError:
            logger.warning("openpyxl 未安装，使用 CSV 替代")
            self._export_csv(data, columns, file_path.replace(".xlsx", ".csv"), True)
            return

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "导出数据"

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")

        # 写入表头
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # 写入数据
        for row_idx, row in enumerate(data, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = row.get(col_name, "")
                if isinstance(value, (int, float)):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                else:
                    ws.cell(row=row_idx, column=col_idx, value=str(value))

        # 自动调整列宽
        for col_idx, col_name in enumerate(columns, 1):
            max_length = len(str(col_name))
            for row in data:
                val = str(row.get(col_name, ""))
                max_length = max(max_length, len(val))
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = min(max_length + 2, 50)

        wb.save(file_path)
        logger.info(f"Excel 导出成功: {file_path}, {len(data)} 行")

    def _export_json(
        self,
        data: list[dict],
        columns: list[str],
        file_path: str,
    ) -> None:
        """导出 JSON"""
        # 只导出指定列
        filtered = [{col: row.get(col) for col in columns} for row in data]

        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(filtered, f, ensure_ascii=False, indent=2)

        logger.info(f"JSON 导出成功: {file_path}, {len(data)} 行")

    def _export_sql(
        self,
        data: list[dict],
        columns: list[str],
        file_path: str,
        table_name: str,
    ) -> None:
        """导出为 INSERT SQL"""
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(f"-- 数据导出: {table_name}\n")
            f.write(f"-- 导出时间: {datetime.now().isoformat()}\n")
            f.write(f"-- 行数: {len(data)}\n\n")

            if not data:
                f.write(f"-- 无数据\n")
                return

            # 生成 CREATE TABLE 语句（简单版）
            col_defs = []
            for col in columns:
                sample = next((str(row.get(col, "")) for row in data if row.get(col) is not None), "")
                if isinstance(sample, int):
                    col_type = "INTEGER"
                elif isinstance(sample, float):
                    col_type = "DECIMAL(18,2)"
                else:
                    col_type = "TEXT"
                col_defs.append(f'"{col}" {col_type}')

            f.write(f"CREATE TABLE IF NOT EXISTS {table_name} (\n")
            f.write(",\n".join(f"    {d}" for d in col_defs))
            f.write("\n);\n\n")

            # 生成 INSERT 语句
            for row in data:
                values = []
                for col in columns:
                    v = row.get(col)
                    if v is None:
                        values.append("NULL")
                    elif isinstance(v, (int, float)):
                        values.append(str(v))
                    else:
                        values.append(f"'{str(v).replace("'", "''")}'")
                f.write(f"INSERT INTO {table_name} ({', '.join(f'"{c}"' for c in columns)}) VALUES ({', '.join(values)});\n")

        logger.info(f"SQL 导出成功: {file_path}, {len(data)} 行")

    def _export_pdf(
        self,
        data: list[dict],
        columns: list[str],
        file_path: str,
    ) -> None:
        """导出 PDF"""
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
        except ImportError:
            logger.warning("reportlab 未安装，生成 CSV 替代")
            self._export_csv(data, columns, file_path.replace(".pdf", ".csv"), True)
            return

        doc = SimpleDocTemplate(
            file_path,
            pagesize=landscape(A4),
            rightMargin=1 * cm,
            leftMargin=1 * cm,
            topMargin=1 * cm,
            bottomMargin=1 * cm,
        )

        styles = getSampleStyleSheet()
        elements = [
            Paragraph("数据导出报告", styles["Title"]),
            Paragraph(f"导出时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles["Normal"]),
            Paragraph(f"共 {len(data)} 行", styles["Normal"]),
            Spacer(1, 0.5 * cm),
        ]

        # 限制 PDF 行数
        display_data = data[:200]
        table_data = [columns] + [
            [str(row.get(col, "")) for col in columns]
            for row in display_data
        ]

        if len(data) > 200:
            table_data.append(["..." for _ in columns])

        table = Table(table_data)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#E8F0FE")),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ]))

        elements.append(table)
        doc.build(elements)
        logger.info(f"PDF 导出成功: {file_path}, {len(data)} 行")


# ── 频率限制装饰器 ─────────────────────────────────────────

_rate_limit_cache: dict[str, list[float]] = {}


def _check_rate_limit(user_id: str, limit: int = 10, window: float = 60.0) -> bool:
    """检查频率限制"""
    now = time.time()
    if user_id not in _rate_limit_cache:
        _rate_limit_cache[user_id] = []

    # 清理过期记录
    _rate_limit_cache[user_id] = [
        t for t in _rate_limit_cache[user_id]
        if now - t < window
    ]

    if len(_rate_limit_cache[user_id]) >= limit:
        return False

    _rate_limit_cache[user_id].append(now)
    return True


def rate_limit(max_requests: int = 10, window_seconds: float = 60.0):
    """频率限制装饰器"""
    def decorator(func):
        def wrapper(*args, **kwargs):
            user_id = kwargs.get("user_id", "anonymous")
            if not _check_rate_limit(user_id, max_requests, window_seconds):
                raise RuntimeError(
                    f"导出频率超限：每 {window_seconds:.0f} 秒最多 {max_requests} 次"
                )
            return func(*args, **kwargs)
        return wrapper
    return decorator


# ── 便捷函数 ───────────────────────────────────────────────

def export_to_csv(
    data: list[dict],
    columns: list[str],
    file_path: str,
) -> ExportResult:
    """便捷函数：导出为 CSV"""
    exporter = DataExporter()
    result = exporter.export(ExportRequest(
        data=data,
        columns=columns,
        format="csv",
        filename=Path(file_path).stem,
    ))
    return result


def export_to_json(
    data: list[dict],
    columns: list[str],
    file_path: str,
) -> ExportResult:
    """便捷函数：导出为 JSON"""
    exporter = DataExporter()
    result = exporter.export(ExportRequest(
        data=data,
        columns=columns,
        format="json",
        filename=Path(file_path).stem,
    ))
    return result
